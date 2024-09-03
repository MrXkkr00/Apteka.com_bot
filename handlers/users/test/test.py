import re

import openpyxl

wb = openpyxl.load_workbook('../Apteka.xlsx')
ws = wb.active
dori_nomi = [ws.cell(row=i, column=1).value for i in range(1, ws.max_row)]
dori_narxi = [ws.cell(row=i, column=2).value for i in range(1, ws.max_row)]


def ignore_case_search_example(text, pattern):
    # text = "I have an Apple and a banana."
    # pattern = r"apple"
    result = re.search(pattern, text, re.IGNORECASE)
    if result:
        return 1
    else:
        return 0



# for i in range(912):
   # javob = ignore_case_search_example(dori, str(dori_nomi[i + 1]))
   #
   # text = "ПАРОДОНТОЛ."
   # pattern = "пар"
   # print(type(pattern))
for i in range(912):
   dori = "тримол"
   if ignore_case_search_example(dori_nomi[i+1], dori) ==1:
      print(dori_nomi[i+1])




            await message.answer(f"💊{i+1}💊{dori_nomi[i + 1]}\n💵 {dori_narxi[i + 1]} $")
    await message.answer(f"Dori raqamini yuboring (faqat raqam)\nMasalan : 44")
    await Searchstate.next()


@dp.message_handler(state=Searchstate.dori)
async def dori_raqami(message: types.Message, state: FSMContext):
    await message.answer(f"Dori qildorini yuboring (faqat raqam)\n Masalan : 4")
    await Searchstate.next()


@dp.message_handler(state=Searchstate.dori)
async def dori_soni(message: types.Message, state: FSMContext):
    await message.answer(f"Siz bilan aloqa uchun cantact yuboring")
    await Searchstate.next()


@dp.message_handler(state=Searchstate.contact)
async def search_cantact(message: types.Message, state: FSMContext):
    await message.answer(f"Sizning buyutmangiz qabul qilindi \n"
                         f"Tez orada siz bilan bo'g'lanamiz\n"
                         f"Yana nimadir hohlasangiz qaytadan buyutma berishingiz mumkin"
                         , reply_markup=bosh_menu)
    await state.finish()
