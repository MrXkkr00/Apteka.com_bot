import openpyxl
import re
from aiogram import types
from aiogram.dispatcher.filters.state import StatesGroup, State
from aiogram.dispatcher import FSMContext
from aiogram.types import ReplyKeyboardRemove, ReplyKeyboardMarkup, \
    KeyboardButton, CallbackQuery

from data.config import ADMINS
from keyboards.default.contact_button import contact_keyboard
from .transliterate import to_cyrillic

from keyboards.default.menu import bosh_menu, qayta_izlash, qayta_izlash_IN
from loader import dp, bot

wb = openpyxl.load_workbook('./data/Apteka.xlsx')
ws = wb.active
dori_nomi = [ws.cell(row=i, column=1).value for i in range(1, ws.max_row)]
dori_narxi = [ws.cell(row=i, column=2).value for i in range(1, ws.max_row)]

izlash_2 = ReplyKeyboardMarkup(keyboard=[
    [
        KeyboardButton(text="🔍Поиск снова"),
    ],
    [
        KeyboardButton(text="Главное меню"),
    ],

], resize_keyboard=True
)


def ignore_case_search_example(text, pattern):
    # text = "I have an Apple and a banana."
    # pattern = r"apple"
    result = re.search(pattern, text, re.IGNORECASE)
    if result:
        return 1
    else:
        return 0


# text = "ПАРОДОНТОЛ."
# pattern = r"пар"
# ignore_case_search_example(text, pattern)


class Searchstate(StatesGroup):
    search = State()
    buyurtma = State()
    dori = State()
    count = State()
    contact = State()


list = []


@dp.message_handler(text="Главное меню", state=[Searchstate.buyurtma, Searchstate.dori, Searchstate.count])
async def search_ct(message: types.Message, state: FSMContext):
    await message.answer(f"Добрый день, добро пожаловать в наш бот", reply_markup=bosh_menu)
    await state.finish()


@dp.message_handler(text="🔍 Поиск лекарств")
async def bot_st1art1(message: types.Message):
    await message.answer(f"Введите название препарата, который вы ищете", reply_markup=ReplyKeyboardRemove())
    await Searchstate.search.set()


@dp.callback_query_handler(text="🔍 Поиск лекарств")
async def buy_co213urses(call: CallbackQuery):
    await call.message.answer(f"Введите название препарата, который вы ищете", reply_markup=ReplyKeyboardRemove())
    await call.message.delete()
    await Searchstate.search.set()



@dp.message_handler(state=Searchstate.search)
async def bot_star1t1(message: types.Message, state: FSMContext):
    msg = message.text
    result = lambda msg: to_cyrillic(msg) if msg.isascii() else msg
    dori = result(msg)
    list = []
    natija = 0
    for i in range(913):
        if ignore_case_search_example(dori_nomi[i + 1], dori) == 1:
            list.append(dori_nomi[i + 1])
            summ = float(dori_narxi[i + 1])
            summa = round(summ, 2)
            natija = 1
            await message.answer(
                f"#️⃣<b>{(i + 1)}</b>         {dori_nomi[i + 1]}\n"
                f"                           💵 {summa} $")

    if natija == 1:
        await message.answer(f"Выберите нужный раздел", reply_markup=qayta_izlash)
        await message.answer(f"Выберите нужный раздел", reply_markup=qayta_izlash_IN)
        await Searchstate.next()

    else:
        await message.answer(f"По вашему запросу ничего не найдено\n"
                             f"Пожалуйста, проверьте и найдите еще раз", reply_markup=bosh_menu)
        await state.finish()


@dp.message_handler(text="🔍Поиск снова", state=[Searchstate.buyurtma, Searchstate.dori, Searchstate.count])
async def bot_st1art1(message: types.Message):
    await message.answer(f"Введите название препарата, который вы ищете", reply_markup=ReplyKeyboardRemove())
    await Searchstate.search.set()


@dp.callback_query_handler(text="📋Разместить заказ", state=Searchstate.buyurtma)
async def buy_courses(call: CallbackQuery):
    await call.message.answer(f"Укажите номер препарата (только номер)\nНапример : 44", reply_markup=izlash_2)
    await call.message.delete()
    await Searchstate.next()


@dp.callback_query_handler(text="🔍Поиск снова", state=Searchstate.buyurtma)
async def buy_courses(call: CallbackQuery):
    await call.message.answer(f"Введите название препарата, который вы ищете", reply_markup=ReplyKeyboardRemove())
    await call.message.delete()
    await Searchstate.search.set()


@dp.callback_query_handler(text="Главное меню", state=Searchstate.buyurtma)
async def buy_courses(call: CallbackQuery, state: FSMContext):
    await call.message.answer(f"Добрый день, добро пожаловать в наш бот", reply_markup=bosh_menu)
    await call.message.delete()
    await state.finish()


@dp.message_handler(text="📋Разместить заказ", state=Searchstate.buyurtma)
async def bot_st1art1(message: types.Message):
    await message.answer(f"Укажите номер препарата (только номер)\nНапример : 44", reply_markup=izlash_2)
    await Searchstate.next()


@dp.message_handler(lambda message: not message.text.isdigit(), state=Searchstate.dori)
async def proc_invalid(message: types.Message):
    """
        If age is invalid
        """
    return await message.reply("Просто введите номер здесь:\nНапример: 125", reply_markup=izlash_2)


@dp.message_handler(lambda message: message.text.isdigit(), state=Searchstate.dori)
async def dori_raqami(message: types.Message, state: FSMContext):
    nomer = int(message.text)
    dori = nomer
    await state.update_data(
        {"dori": dori}
    )
    await message.answer(f"Пожалуйста введите количество лекарства (только в цифрах)\n Например : 4",
                         reply_markup=izlash_2)
    await Searchstate.next()


@dp.message_handler(lambda message: not message.text.isdigit(), state=Searchstate.count)
async def proc_inva123lid(message: types.Message):
    """
        If age is invalid
        """
    return await message.reply("Просто введите номер здесь:\nНапример: 5", reply_markup=izlash_2)


@dp.message_handler(lambda message: message.text.isdigit(), state=Searchstate.count)
async def dori_soni(message: types.Message, state: FSMContext):
    nomer1 = int(message.text)
    await state.update_data(
        {"count": nomer1}
    )
    await message.answer(f"Отправьте контакт, чтобы связаться с вами", reply_markup=contact_keyboard)
    await Searchstate.next()


@dp.message_handler(content_types='contact', state=Searchstate.contact)
async def search_cantact(message: types.Message, state: FSMContext):
    contact = message.contact
    data = await state.get_data()
    dori = data.get("dori")
    count = data.get("count")
    summ = float(count) * float(dori_narxi[dori])
    summa = round(summ, 2)
    # 🙎🏻✅💊📲
    await message.answer(f"🙎🏻‍♂Клиент                 @{message.from_user.username}\n"
                         f"💊 Товары              {dori_nomi[dori]}\n"
                         f"✅ Количество      {count}\n"
                         f"💵 Общая сумма   {summa} $ \n"
                         f"📲 Контакт            T.me/{contact.phone_number}")
    await bot.send_message(ADMINS[1], f"🙎🏻‍♂Клиент                 @{message.from_user.username}\n"
                                      f"💊 Товары              {dori_nomi[dori]}\n"
                                      f"✅ Количество      {count}\n"
                                      f"💵 Общая сумма   {summa} $ \n"
                                      f"📲 Контакт            T.me/{contact.phone_number}")
    await message.answer(f"Ваш заказ принят\n"
                         f"Мы скоро свяжемся с вами\n"
                         f"Если вы хотите что-то еще, вы можете заказать еще раз"
                         , reply_markup=bosh_menu)
    await state.finish()
